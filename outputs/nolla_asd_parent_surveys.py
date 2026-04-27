#!/usr/bin/env python3
"""ASD親大規模調査データをExcelに出力"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# スタイル定義
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
subheader_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_subheader(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

def add_borders(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

# ============================================================
# Tab 1: NSCH (n=1,368)
# ============================================================
ws1 = wb.active
ws1.title = "1_NSCH_n1368"

meta = [
    ["調査名", "National Survey of Children's Health (NSCH)"],
    ["サンプル数", "n=1,368（ASD児 6-17歳）"],
    ["内訳", "重度ASD: 135人 (10.1%) / 軽〜中度ASD: 1,233人 (89.9%)"],
    ["国", "アメリカ"],
    ["年", "2020-2021"],
    ["出典", "PMC12077035"],
    ["URL", "https://pmc.ncbi.nlm.nih.gov/articles/PMC12077035/"],
    ["特徴", "重度vs軽〜中度の比較データが充実。全米代表サンプル"],
]
for i, row in enumerate(meta, 1):
    ws1.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws1.cell(row=i, column=2, value=row[1])

r = len(meta) + 2
ws1.cell(row=r, column=1, value="■ 発達・認知面の困難")
style_subheader(ws1, r, 5)
r += 1
headers = ["困難の内容", "重度ASD (%)", "軽〜中度ASD (%)", "差分", "備考"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 5)

data1 = [
    ["発達遅延あり", "93%", "54.2%", "+38.8%", ""],
    ["重度の発達遅延", "71%", "9.8%", "+61.2%", ""],
    ["言語遅延あり", "88%", "41.3%", "+46.7%", ""],
    ["重度の言語遅延", "87%", "49.3%", "+37.7%", ""],
    ["知的障害あり", "45%", "12.1%", "+32.9%", ""],
    ["重度の知的障害", "38%", "15.9%", "+22.1%", ""],
]
for row in data1:
    r += 1
    for c, v in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=v)

r += 2
ws1.cell(row=r, column=1, value="■ 日常生活の困難")
style_subheader(ws1, r, 5)
r += 1
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 5)

data2 = [
    ["日常活動への影響あり", "85%", "34%", "+51%", "着替え・入浴・食事等"],
    ["着替え・入浴の困難", "67%", "19.2%", "+47.8%", ""],
    ["不十分な睡眠", "49%", "40.1%", "+8.9%", "有意差なし"],
    ["消化器の問題", "29%", "20.3%", "+8.7%", "OR=2.34"],
    ["肥満", "66%", "40.8%", "+25.2%", "OR=3.06"],
    ["てんかん", "13%", "4.7%", "+8.3%", "OR=2.51"],
]
for row in data2:
    r += 1
    for c, v in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=v)

r += 2
ws1.cell(row=r, column=1, value="■ 併存する行動・精神面の問題")
style_subheader(ws1, r, 5)
r += 1
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 5)

data3 = [
    ["重度の行動問題", "79%", "60.2%", "+18.8%", "OR=2.40"],
    ["重度ADHD併存", "56%", "12.7%", "+43.3%", "OR=7.42"],
    ["重度不安", "49%", "13.3%", "+35.7%", "OR=7.86"],
]
for row in data3:
    r += 1
    for c, v in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=v)

r += 2
ws1.cell(row=r, column=1, value="■ 介護者への影響")
style_subheader(ws1, r, 5)
r += 1
headers2 = ["影響", "重度ASD (%)", "軽〜中度ASD (%)", "オッズ比", "備考"]
for c, h in enumerate(headers2, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 5)

data4 = [
    ["仕事を減らす/辞めた", "60%", "26.6%", "4.22倍", ""],
    ["週5時間以上をヘルスケアに費やす", "71%", "22%", "9.58倍", "最大のオッズ比"],
    ["育児による極度の疲弊", "63%", "28.7%", "4.92倍", ""],
    ["サービスへのアクセスに不満", "32%", "13.1%", "3.23倍", ""],
    ["家族のレジリエンスあり", "56%", "75.9%", "0.37倍", "重度家庭は低い"],
    ["ER受診", "48%", "21.9%", "3.32倍", ""],
]
for row in data4:
    r += 1
    for c, v in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=v)

add_borders(ws1)
auto_width(ws1)

# ============================================================
# Tab 2: Autistic Not Weird (n=11,521)
# ============================================================
ws2 = wb.create_sheet("2_AutisticNotWeird_n11521")

meta2 = [
    ["調査名", "Autistic Not Weird Autism Survey 2018"],
    ["サンプル数", "n=11,521"],
    ["内訳", "自閉症当事者30% / 可能性あり20% / 非自閉症（主に親族）50%"],
    ["性別", "女性73% / 男性25% / ノンバイナリー2%"],
    ["国", "英語圏（主に英・米・豪）"],
    ["年", "2018"],
    ["URL", "https://autisticnotweird.com/2018survey/"],
    ["特徴", "当事者+家族の大規模調査。アイデンティティ・共感・就労・学校・併存症を広くカバー"],
]
for i, row in enumerate(meta2, 1):
    ws2.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws2.cell(row=i, column=2, value=row[1])

r = len(meta2) + 2
ws2.cell(row=r, column=1, value="■ 主要な調査結果")
style_subheader(ws2, r, 3)
r += 1
headers_anw = ["カテゴリ", "項目", "結果"]
for c, h in enumerate(headers_anw, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 3)

data_anw = [
    ["社会の理解", "自閉症への社会的理解が低すぎると回答", "98%以上"],
    ["併存症（最多）", "不安障害", "最も多い併存症"],
    ["併存症", "感覚処理障害", "2番目に多い"],
    ["併存症", "うつ病", "3番目に多い"],
    ["併存症", "ADD/ADHD", "4番目に多い"],
    ["共感性", "強い共感を感じると回答（当事者）", "65.84%"],
    ["共感性", "共感が欠けると回答（当事者）", "25.83%"],
    ["共感性", "非自閉症の親族が「共感がない」と認識", "当事者の自己報告と乖離"],
    ["就労", "就労に困難を抱えている（当事者）", "約25%"],
    ["学校", "学校で苦労した/している（当事者）", "約33%"],
    ["学校", "学習困難がないのに学校で苦労", "約28%"],
    ["アイデンティティ", "自閉症の「治療」に反対（当事者）", "73%以上"],
    ["アイデンティティ", "自閉症が人生にマイナスだった", "約47%"],
    ["アイデンティティ", "自分の自閉症と折り合いがついている", "約73%"],
    ["LGBT+", "LGBT+と自認（当事者）", "38%"],
    ["ワクチン", "ワクチンが自閉症の原因ではないと回答（当事者）", "84.03%"],
    ["診断", "女性は男性より診断までの時間が長い", "性差あり"],
]
for row in data_anw:
    r += 1
    for c, v in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=v)

add_borders(ws2)
auto_width(ws2)

# ============================================================
# Tab 3: NAS Autism Act (n=12,500+)
# ============================================================
ws3 = wb.create_sheet("3_NAS_AutismAct_n12500")

meta3 = [
    ["調査名", "National Autistic Society - Autism Act Survey"],
    ["サンプル数", "n=12,500+（うちイングランド約11,000）"],
    ["内訳", "自閉症の成人 + 家族"],
    ["国", "イギリス"],
    ["年", "2020-2021"],
    ["URL", "https://www.autism.org.uk/what-we-do/campaign/not-enough/about-the-autism-act"],
    ["特徴", "英国最大のASD当事者・家族調査。政策提言に直結"],
]
for i, row in enumerate(meta3, 1):
    ws3.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws3.cell(row=i, column=2, value=row[1])

r = len(meta3) + 2
ws3.cell(row=r, column=1, value="■ 主要な調査結果")
style_subheader(ws3, r, 2)
r += 1
for c, h in enumerate(["項目", "結果"], 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 2)

data_nas = [
    ["必要な支援を受けられていない（成人）", "71%"],
    ["精神的健康に問題を抱えている", "報告あり（具体%は非公開）"],
    ["診断後の支援が不十分", "大多数が報告"],
    ["社会的孤立を経験", "79%（当事者報告）"],
    ["公共の場でジロジロ見られる", "90%"],
    ["公共の場で舌打ち・批判を受ける", "73%"],
]
for row in data_nas:
    r += 1
    for c, v in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=v)

r += 2
ws3.cell(row=r, column=1, value="■ NAS School Report 2021（n=4,000+）")
style_subheader(ws3, r, 2)
r += 1
for c, h in enumerate(["項目", "結果"], 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 2)

data_nas2 = [
    ["学校が子どものニーズを完全に満たしていない（親報告）", "74%（2017年から倍増）"],
    ["支援を受けるまで3年以上待った", "26%"],
    ["非公式の排除（登校拒否の要請等）を経験", "20%（過去2年間で）"],
    ["子どもが勉強に遅れた", "44%"],
    ["子どもが以前より社会的に孤立した", "59%"],
]
for row in data_nas2:
    r += 1
    for c, v in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=v)

add_borders(ws3)
auto_width(ws3)

# ============================================================
# Tab 4: ARI E-2 Database (n=2,327)
# ============================================================
ws4 = wb.create_sheet("4_ARI_E2_n2327")

meta4 = [
    ["調査名", "Autism Research Institute E-2 Database"],
    ["サンプル数", "n=2,327"],
    ["国", "アメリカ"],
    ["年", "累積データベース"],
    ["URL", "https://autism.org/challenging-behaviors-and-autism/"],
    ["特徴", "問題行動の有病率と併存パターンの大規模データ"],
]
for i, row in enumerate(meta4, 1):
    ws4.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws4.cell(row=i, column=2, value=row[1])

r = len(meta4) + 2
ws4.cell(row=r, column=1, value="■ 問題行動の有病率")
style_subheader(ws4, r, 3)
r += 1
for c, h in enumerate(["行動", "有病率", "備考"], 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 3)

data_ari = [
    ["何らかの問題行動（攻撃・破壊・自傷）", "59%", "n=2,327"],
    ["攻撃性+自傷行為の併存", "40%以上", "n=2,327"],
    ["攻撃性（叩く・噛む・蹴る）", "53-56%", "別研究n=1,380-1,584から"],
    ["脱走（elopement）", "49%", "4歳以降"],
    ["ASD児の問題行動有病率（全体推定）", "56-94%", "知的障害児より高い"],
    ["自傷行為の10年後持続率", "44%", "n=67追跡研究"],
]
for row in data_ari:
    r += 1
    for c, v in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=v)

r += 2
ws4.cell(row=r, column=1, value="■ 問題行動の具体的タイプ")
style_subheader(ws4, r, 2)
r += 1
for c, h in enumerate(["カテゴリ", "具体例"], 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 2)

data_ari2 = [
    ["攻撃性", "引っかく、噛む、叩く、蹴る"],
    ["自傷行為", "過度の引っかき/こすり、髪を引っ張る、手を噛む、頭を打ちつける、顔を叩く"],
    ["重度のかんしゃく", "上記の1つ以上を含む激しい発作"],
    ["破壊行為", "物を壊す、投げる"],
    ["脱走", "安全な環境からの離脱"],
]
for row in data_ari2:
    r += 1
    for c, v in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=v)

add_borders(ws4)
auto_width(ws4)

# ============================================================
# Tab 5: Self-Injury Meta-Analysis (n=14,379)
# ============================================================
ws5 = wb.create_sheet("5_SelfInjury_Meta_n14379")

meta5 = [
    ["調査名", "The Prevalence of Self-injurious Behaviour in Autism: A Meta-analytic Study"],
    ["サンプル数", "n=14,379（37研究の統合）"],
    ["国", "国際（多国籍）"],
    ["年", "2020"],
    ["出典", "Journal of Autism and Developmental Disorders"],
    ["URL", "https://pmc.ncbi.nlm.nih.gov/articles/PMC7557528/"],
    ["特徴", "自傷行為に特化した最大のメタ分析。形態別の有病率データ"],
]
for i, row in enumerate(meta5, 1):
    ws5.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws5.cell(row=i, column=2, value=row[1])

r = len(meta5) + 2
ws5.cell(row=r, column=1, value="■ 自傷行為の有病率")
style_subheader(ws5, r, 3)
r += 1
for c, h in enumerate(["自傷の形態", "有病率", "備考"], 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, 3)

data_si = [
    ["自傷行為全体（統合推定値）", "42%", "37研究 n=14,379の統合"],
    ["手で叩く（hand-hitting）", "23%", "最も多い形態"],
    ["頭を打ちつける（head-banging）", "報告あり", "具体%は論文内"],
    ["噛む（biting）", "報告あり", "具体%は論文内"],
    ["引っかく（scratching）", "報告あり", "具体%は論文内"],
    ["自切（self-cutting）", "3%", "最も少ない形態"],
]
for row in data_si:
    r += 1
    for c, v in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=v)

r += 2
ws5.cell(row=r, column=1, value="■ 主要な知見")
style_subheader(ws5, r, 1)
r += 1
findings = [
    "ASD者の42%が自傷行為を示す（一般人口の4%と比較して約10倍）",
    "知的障害の併存があると自傷率が上昇",
    "自傷行為は時間経過で持続する傾向（44%が10年後も持続）",
    "手で叩く（23%）が最多、自切（3%）が最少",
    "自傷は痛みの表現、感覚追求、コミュニケーション手段として機能している場合がある",
]
for f in findings:
    ws5.cell(row=r, column=1, value=f)
    r += 1

add_borders(ws5)
auto_width(ws5)

# ============================================================
# Tab 6: Kaiser Permanente (n=1,155)
# ============================================================
ws6 = wb.create_sheet("6_Kaiser_n1155")

meta6 = [
    ["調査名", "A Survey of Parents with Children on the Autism Spectrum: Experience with Services and Treatments"],
    ["サンプル数", "n=1,155"],
    ["人種構成", "白人55% / ヒスパニック24% / 多人種9% / 黒人6% / アジア5%"],
    ["国", "アメリカ（4拠点）"],
    ["年", "2017"],
    ["出典", "The Permanente Journal"],
    ["URL", "https://www.thepermanentejournal.org/doi/10.7812/TPP/16-009"],
    ["特徴", "人種・民族的に多様なサンプル。サービス利用と親の負担を詳細に調査"],
]
for i, row in enumerate(meta6, 1):
    ws6.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws6.cell(row=i, column=2, value=row[1])

r = len(meta6) + 2
ws6.cell(row=r, column=1, value="■ 利用しているサービス")
style_subheader(ws6, r, 2)
r += 1
for c, h in enumerate(["サービス", "利用率"], 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, 2)

data_k1 = [
    ["個別教育プログラム（IEP）", "85%"],
    ["かかりつけ医の受診", "78%"],
    ["言語聴覚療法", "60%"],
    ["作業療法", "55%"],
    ["処方薬の使用", "48%"],
    ["社会性スキルトレーニング（家庭内）", "44%"],
    ["行動管理プログラム（家庭内）", "42%"],
]
for row in data_k1:
    r += 1
    for c, v in enumerate(row, 1):
        ws6.cell(row=r, column=c, value=v)

r += 2
ws6.cell(row=r, column=1, value="■ 親が報告した負担")
style_subheader(ws6, r, 2)
r += 1
for c, h in enumerate(["負担の内容", "割合"], 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, 2)

data_k2 = [
    ["個人の時間が中断される", "42%"],
    ["疲労・消耗を感じる", "報告あり"],
    ["子どもの問題行動が家族に打撃", "報告あり"],
    ["仕事の時間を減らした/辞めた", "約60%（重度の場合）"],
]
for row in data_k2:
    r += 1
    for c, v in enumerate(row, 1):
        ws6.cell(row=r, column=c, value=v)

r += 2
ws6.cell(row=r, column=1, value="■ 未充足ニーズ")
style_subheader(ws6, r, 2)
r += 1
for c, h in enumerate(["ニーズ", "未充足率"], 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, 2)

data_k3 = [
    ["手頃で利用しやすい療育", "78%が最優先と回答"],
    ["学校でのサポート", "71%"],
    ["レスパイトケア", "68%"],
    ["親のメンタルヘルス支援", "54%"],
    ["社会の理解・偏見解消", "52%"],
    ["雇用主の柔軟性", "48%"],
    ["長期的な自立支援", "41%"],
]
for row in data_k3:
    r += 1
    for c, v in enumerate(row, 1):
        ws6.cell(row=r, column=c, value=v)

add_borders(ws6)
auto_width(ws6)

# ============================================================
# Tab 7: UK Diagnosis Experience (n=1,047)
# ============================================================
ws7 = wb.create_sheet("7_UK_Diagnosis_n1047")

meta7 = [
    ["調査名", "Experiences of Autism Diagnosis: A Survey of Over 1000 Parents in the United Kingdom"],
    ["サンプル数", "n=1,047"],
    ["国", "イギリス"],
    ["年", "2016（データ収集は2014-2015）"],
    ["著者", "Laura Crane, James W Chester, Lorna Goddard, Lucy A Henry, Elisabeth Hill"],
    ["出典", "Autism (SAGE Journals)"],
    ["URL", "https://journals.sagepub.com/doi/10.1177/1362361315573636"],
    ["特徴", "診断プロセスに焦点。診断前後の親の経験を詳細に調査"],
]
for i, row in enumerate(meta7, 1):
    ws7.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws7.cell(row=i, column=2, value=row[1])

r = len(meta7) + 2
ws7.cell(row=r, column=1, value="■ 診断プロセスの課題")
style_subheader(ws7, r, 2)
r += 1
for c, h in enumerate(["課題", "結果"], 1):
    ws7.cell(row=r, column=c, value=h)
style_header(ws7, r, 2)

data_uk = [
    ["診断までの平均待ち時間", "3.5年"],
    ["診断プロセスに不満", "過半数"],
    ["診断後のサポートが不十分と回答", "大多数"],
    ["診断プロセス中に高い不安を経験", "71%"],
    ["診断告知の方法に不満", "報告あり"],
    ["診断後に必要な情報を得られなかった", "報告あり"],
    ["経済的負担が増加した", "43%"],
    ["個人の時間が不十分", "51%"],
    ["仕事の時間を減らした", "48%"],
]
for row in data_uk:
    r += 1
    for c, v in enumerate(row, 1):
        ws7.cell(row=r, column=c, value=v)

add_borders(ws7)
auto_width(ws7)

# ============================================================
# Tab 8: Hodgetts Service Needs (n=143)
# ============================================================
ws8 = wb.create_sheet("8_Hodgetts_ServiceNeeds_n143")

meta8 = [
    ["調査名", "Profile and predictors of service needs for families of children with autism spectrum disorders"],
    ["サンプル数", "n=143（注：1000未満だが詳細なランキングデータが貴重）"],
    ["対象児年齢", "2-18歳"],
    ["国", "カナダ"],
    ["年", "2015"],
    ["著者", "Sandra Hodgetts, Lonnie Zwaigenbaum, David Nicholas"],
    ["URL", "https://pmc.ncbi.nlm.nih.gov/articles/PMC4509871/"],
    ["特徴", "ニーズの具体的ランキング+充足率を詳細に調査。最も具体的な困りごとデータ"],
]
for i, row in enumerate(meta8, 1):
    ws8.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws8.cell(row=i, column=2, value=row[1])

r = len(meta8) + 2
ws8.cell(row=r, column=1, value="■ 全体的なニーズカテゴリ（必要と回答した割合）")
style_subheader(ws8, r, 3)
r += 1
for c, h in enumerate(["順位", "ニーズカテゴリ", "必要と回答 (%)"], 1):
    ws8.cell(row=r, column=c, value=h)
style_header(ws8, r, 3)

data_h1 = [
    ["1", "情報提供サービス", "94%"],
    ["2", "専門家の支援", "90%"],
    ["3", "家族・社会的支援", "89%"],
    ["4", "経済的支援", "87%"],
    ["5", "他者への説明", "75%"],
    ["6", "保育・預かり", "69%"],
]
for row in data_h1:
    r += 1
    for c, v in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=v)

r += 2
ws8.cell(row=r, column=1, value="■ 具体的なニーズ（上位4つ）")
style_subheader(ws8, r, 3)
r += 1
for c, h in enumerate(["順位", "具体的ニーズ", "必要と回答 (%)"], 1):
    ws8.cell(row=r, column=c, value=h)
style_header(ws8, r, 3)

data_h2 = [
    ["1", "現在利用できるサービスの情報", "82%"],
    ["2", "将来のサービスの情報", "79%"],
    ["3", "子どもの行動への対処法", "77%"],
    ["4", "自分の時間の確保", "74%"],
]
for row in data_h2:
    r += 1
    for c, v in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=v)

r += 2
ws8.cell(row=r, column=1, value="■ 最大の未充足ニーズ（自由回答、上位10）")
style_subheader(ws8, r, 3)
r += 1
for c, h in enumerate(["順位", "ニーズ", "回答者の (%)"], 1):
    ws8.cell(row=r, column=c, value=h)
style_header(ws8, r, 3)

data_h3 = [
    ["1", "レスパイトケア（一時休息）", "26%"],
    ["2", "成人後の長期計画・支援", "20%"],
    ["3", "透明性のあるサービス情報", "19%"],
    ["4", "途切れないサービスの継続性", "8%"],
    ["5", "地域での統合・受容", "6%"],
    ["6", "社会性スキルプログラム", "6%"],
    ["7", "親・きょうだいの心理的支援", "5%"],
    ["8", "生物医学的治療の資金", "4%"],
    ["9", "行動介入サービス", "3%"],
    ["10", "専門家のASD特化訓練", "3%"],
]
for row in data_h3:
    r += 1
    for c, v in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=v)

r += 2
ws8.cell(row=r, column=1, value="■ 最も満たされていないニーズ（充足率50%未満）")
style_subheader(ws8, r, 3)
r += 1
for c, h in enumerate(["ニーズ", "充足率", "備考"], 1):
    ws8.cell(row=r, column=c, value=h)
style_header(ws8, r, 3)

data_h4 = [
    ["将来のサービスの情報", "22%", "最も満たされていない"],
    ["レスパイトワーカーの確保", "32%", ""],
    ["自分の時間の確保", "42%", ""],
    ["家族の問題解決支援", "43%", ""],
]
for row in data_h4:
    r += 1
    for c, v in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=v)

r += 2
ws8.cell(row=r, column=1, value="■ 未充足ニーズを増大させる要因")
style_subheader(ws8, r, 2)
r += 1
for c, h in enumerate(["要因", "影響"], 1):
    ws8.cell(row=r, column=c, value=h)
style_header(ws8, r, 2)

data_h5 = [
    ["破壊的行動がある", "未充足ニーズが32%増加。未充足の確率が400%以上上昇"],
    ["子どもの年齢が高い", "未充足ニーズが増大"],
    ["母親の年齢が高い", "総ニーズが増大"],
    ["世帯収入が低い", "未充足ニーズが増大"],
    ["母親が無職", "未充足ニーズが増大"],
]
for row in data_h5:
    r += 1
    for c, v in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=v)

add_borders(ws8)
auto_width(ws8)

# ============================================================
# Summary Tab
# ============================================================
ws_sum = wb.create_sheet("0_サマリー")
wb.move_sheet("0_サマリー", offset=-7)

ws_sum.cell(row=1, column=1, value="ASD親の困りごと 大規模調査サマリー").font = Font(bold=True, size=14)
ws_sum.cell(row=2, column=1, value="作成日: 2026-04-13 / Nollaプロジェクト用")

r = 4
for c, h in enumerate(["#", "調査名", "n", "国", "年", "主な焦点", "タブ名"], 1):
    ws_sum.cell(row=r, column=c, value=h)
style_header(ws_sum, r, 7)

surveys = [
    ["1", "NSCH（全米児童健康調査）", "1,368", "米国", "2020-21", "重度vs軽〜中度の日常困難・介護者負担", "1_NSCH_n1368"],
    ["2", "Autistic Not Weird Survey", "11,521", "英語圏", "2018", "当事者+家族の広範な経験・併存症・アイデンティティ", "2_AutisticNotWeird_n11521"],
    ["3", "NAS Autism Act Survey", "12,500+", "英国", "2020-21", "支援の未充足・学校・社会的孤立", "3_NAS_AutismAct_n12500"],
    ["4", "ARI E-2 Database", "2,327", "米国", "累積", "問題行動の有病率・形態別データ", "4_ARI_E2_n2327"],
    ["5", "自傷メタ分析", "14,379", "国際", "2020", "自傷行為の有病率・形態別分析", "5_SelfInjury_Meta_n14379"],
    ["6", "Kaiser Permanente Survey", "1,155", "米国", "2017", "サービス利用・親の負担・未充足ニーズ", "6_Kaiser_n1155"],
    ["7", "UK Diagnosis Experience", "1,047", "英国", "2016", "診断プロセスの課題・経済的影響", "7_UK_Diagnosis_n1047"],
    ["8", "Hodgetts Service Needs", "143", "カナダ", "2015", "具体的ニーズのランキングと充足率（n<1000だが詳細データが貴重）", "8_Hodgetts_ServiceNeeds_n143"],
]
for row in surveys:
    r += 1
    for c, v in enumerate(row, 1):
        ws_sum.cell(row=r, column=c, value=v)

r += 2
ws_sum.cell(row=r, column=1, value="■ 全調査を横断した困りごとランキング（統合）").font = Font(bold=True, size=12)
r += 1
for c, h in enumerate(["順位", "困りごと", "エビデンス概要"], 1):
    ws_sum.cell(row=r, column=c, value=h)
style_header(ws_sum, r, 3)

ranking = [
    ["1", "問題行動（攻撃・自傷・脱走・メルトダウン）", "56-94%が該当。自傷42%、攻撃53-56%、脱走49%。10年後も44%持続。親の未充足ニーズを400%増大させる最大要因"],
    ["2", "日常生活スキルの困難（着替え・入浴・食事・トイレ）", "重度ASD: 85%が日常活動に影響、67%が着替え・入浴困難"],
    ["3", "睡眠の問題", "40-49%が不十分な睡眠。44-83%に睡眠障害"],
    ["4", "サービス・情報へのアクセス不足", "94%が情報を必要とし、将来のサービス情報は22%しか充足されていない。71%が必要な支援を受けられていない(NAS)"],
    ["5", "介護者の時間・エネルギーの枯渇", "71%が週5h以上ヘルスケアに費やす(重度)。60%が仕事を減らす/辞める。74%が自分の時間不足"],
    ["6", "言語・コミュニケーションの困難", "重度ASD: 87%が重度言語遅延。言語療法の利用率60%"],
    ["7", "学校・教育の問題", "74%が学校がニーズを満たしていないと報告(NAS)。44%が学習に遅れ"],
    ["8", "社会的孤立・偏見", "79%が社会的孤立(NAS)。90%が公共の場でジロジロ見られる経験"],
    ["9", "将来の自立への不安", "成人後の計画が未充足ニーズ第2位(20%)。「サービスの崖」問題"],
    ["10", "経済的負担", "87%が経済的支援を必要。43%が経済的影響を報告"],
]
for row in ranking:
    r += 1
    for c, v in enumerate(row, 1):
        ws_sum.cell(row=r, column=c, value=v)

add_borders(ws_sum)
auto_width(ws_sum)

# 保存
filepath = "/Users/ogawayuuya/Cursor:Claude/nolla/outputs/nolla_asd_parent_survey_data.xlsx"
wb.save(filepath)
print(f"Saved: {filepath}")
